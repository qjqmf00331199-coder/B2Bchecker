"""F-06: 조회 결과 테이블을 엑셀(.xlsx)로 내보내기 / 엑셀에서 데이터 가져오기."""
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QFileDialog, QMessageBox, QTableWidget


def export_table_to_excel(table: QTableWidget, parent=None, default_filename: str = "export.xlsx") -> None:
    if table.rowCount() == 0:
        QMessageBox.information(parent, "엑셀 내보내기", "내보낼 데이터가 없습니다.")
        return

    file_path, _ = QFileDialog.getSaveFileName(
        parent, "엑셀로 저장", default_filename, "Excel Files (*.xlsx)"
    )
    if not file_path:
        return
    if not file_path.lower().endswith(".xlsx"):
        file_path += ".xlsx"

    workbook = Workbook()
    sheet = workbook.active

    headers = [table.horizontalHeaderItem(col).text() for col in range(table.columnCount())]
    sheet.append(headers)

    for row in range(table.rowCount()):
        values = []
        for col in range(table.columnCount()):
            item = table.item(row, col)
            values.append(item.text() if item is not None else "")
        sheet.append(values)

    try:
        workbook.save(file_path)
    except OSError as exc:
        QMessageBox.critical(parent, "저장 실패", f"파일 저장 중 오류가 발생했습니다:\n{exc}")
        return

    QMessageBox.information(parent, "엑셀 내보내기", f"저장되었습니다:\n{file_path}")


def import_rows_from_excel(parent=None, dialog_title: str = "엑셀에서 가져오기") -> Optional[List[tuple]]:
    """첫 번째 시트를 읽어 헤더 행을 제외한 데이터 행을 반환한다. 취소/실패 시 None."""
    file_path, _ = QFileDialog.getOpenFileName(parent, dialog_title, "", "Excel Files (*.xlsx)")
    if not file_path:
        return None

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except (OSError, KeyError) as exc:
        QMessageBox.critical(parent, "가져오기 실패", f"파일을 여는 중 오류가 발생했습니다:\n{exc}")
        return None

    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    return rows
